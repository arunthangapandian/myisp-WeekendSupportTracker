"""
Weekend Production Support Tracker - Flask Backend
Equivalent of server.js, runs locally without any cloud hosting.

Install dependencies:
    pip install flask flask-cors openpyxl

Run:
    python app_Flask.py
"""

import json
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory, abort
from flask_cors import CORS
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

app = Flask(__name__)
CORS(app)

PORT = int(os.environ.get("PORT", 5000))

# ── File paths ───────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / "data.json"
UPLOADS_DIR = BASE_DIR / "uploads"
FRONTEND_BUILD = BASE_DIR.parent / "frontend" / "build"
UPLOADS_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv", "pdf"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB

# ── In-memory store ──────────────────────────────────────────
entries = {}
deleted_items = []
changelogs = {}
employees = []
resource_upload_history = []

# ── Persistent save ──────────────────────────────────────────
def save_data():
    """Save all data to local JSON file."""
    payload = {
        "entries": entries,
        "deletedItems": deleted_items,
        "employees": employees,
        "changelogs": changelogs,
        "resourceUploadHistory": resource_upload_history,
    }
    try:
        DATA_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"Failed to save data.json: {e}")


def load_data_from_file():
    """Load data from local JSON file."""
    try:
        if DATA_FILE.exists():
            raw = DATA_FILE.read_text(encoding="utf-8")
            parsed = json.loads(raw)
            return {
                "entries": parsed.get("entries", {}),
                "deletedItems": parsed.get("deletedItems", []),
                "employees": parsed.get("employees") or [],
                "changelogs": parsed.get("changelogs", {}),
                "resourceUploadHistory": parsed.get("resourceUploadHistory", []),
            }
    except Exception as e:
        print(f"Failed to load data.json: {e}")
    return None


# ── Static data ──────────────────────────────────────────────
RELEASE_OWNERS = [
    {"id": "arun.thangapandian", "name": "Arun Thangapandian"},
    {"id": "priya.sharma", "name": "Priya Sharma"},
    {"id": "rahul.verma", "name": "Rahul Verma"},
    {"id": "sneha.patel", "name": "Sneha Patel"},
    {"id": "vikram.singh", "name": "Vikram Singh"},
]

CAREER_LEVELS = [
    "Analyst", "Senior Analyst", "Consultant",
    "Manager", "Senior Manager", "Associate Director",
]

HARDCODED_CAREER_LEVELS = {
    "shini.vv": 9,
    "vishnu.ramalingam": 9,
    "srinivasan.selvam": 7,
    "d.sampathkumar": 8,
}

# ── Helpers ──────────────────────────────────────────────────
def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())


def file_timestamp():
    d = datetime.now()
    return d.strftime("%Y-%m-%d_%H-%M")


def title_case(s):
    return re.sub(r"\b\w", lambda m: m.group().upper(), s or "")


def parse_12h(s):
    m = re.match(r"^(\d{1,2})(?::(\d{2}))?\s*(AM|PM)$", s.strip(), re.IGNORECASE)
    if not m:
        return float("nan")
    h = int(m.group(1))
    mins = int(m.group(2)) if m.group(2) else 0
    period = m.group(3).upper()
    if h == 12:
        h = 0
    if period == "PM":
        h += 12
    return h + mins / 60


def calc_total_hours(time_str):
    if not time_str or "-" not in time_str:
        return ""
    parts = [p.strip() for p in time_str.split("-", 1)]
    if len(parts) != 2:
        return ""
    start = parse_12h(parts[0])
    end = parse_12h(parts[1])
    import math
    if math.isnan(start) or math.isnan(end):
        return ""
    diff = end - start
    if diff <= 0:
        diff += 24
    return f"{int(diff)} hrs" if diff % 1 == 0 else f"{diff:.1f} hrs"


def get_support_type(hrs_str):
    if not hrs_str:
        return ""
    try:
        num = float(hrs_str.split()[0])
    except (ValueError, IndexError):
        return ""
    if num >= 6.5:
        return "On Call Primary Support"
    if num >= 4.5:
        return "On Call Secondary Support"
    return ""


def get_numeric_level(line_item):
    """Get numeric level - prioritize level field, fallback to parsing careerLevel"""
    # First, check if there's a numeric level field
    if line_item.get("level") is not None:
        return line_item["level"]
    # Fallback: try to parse careerLevel string
    career_level = line_item.get("careerLevel", "")
    if not career_level:
        return "—"
    # If it starts with "Level", extract the number
    if "level" in career_level.lower():
        num_str = re.sub(r"[^0-9]", "", str(career_level))
        try:
            return int(num_str)
        except ValueError:
            return "—"
    # Map common abbreviations to numeric levels
    level_map = {
        "analyst": 10,
        "senior analyst": 9,
        "se": 9,
        "consultant": 8,
        "sse": 10,
        "manager": 7,
        "senior manager": 7,
        "associate director": 7,
    }
    lower = career_level.lower().strip()
    return level_map.get(lower, career_level)


def lookup_employee_info(line_item):
    """Lookup employee from resource list by name and return empId and level"""
    if not line_item.get("name"):
        return {"empId": "", "level": None}
    
    # If empId already exists and level exists, return them
    if line_item.get("empId") and line_item.get("level") is not None:
        return {"empId": line_item["empId"], "level": line_item["level"]}
    
    # Lookup in employees list
    name_lower = line_item["name"].lower().strip()
    emp = next((e for e in employees if 
                e.get("name") and e["name"].lower().strip() == name_lower or
                e.get("id") and e["id"].lower().strip() == name_lower), None)
    
    if emp:
        return {
            "empId": emp.get("id", ""),
            "level": emp.get("level") if emp.get("level") is not None else None
        }
    
    # Return existing values or defaults
    return {
        "empId": line_item.get("empId", ""),
        "level": line_item.get("level") if line_item.get("level") is not None else None
    }


def allowed_file(filename):
    ext = Path(filename).suffix.lower().lstrip(".")
    return ext in ALLOWED_EXTENSIONS


# ── Excel styling ─────────────────────────────────────────────
HEADER_BG = "FF3730A3"  # ARGB (openpyxl uses ARGB without #)
WHITE = "FFFFFFFF"
BLACK = "FF000000"
DUPLICATE_BG = "FFFFFF00"

thin_side = Side(style="thin", color="FF000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def style_sheet(ws, columns, row_data_list, get_row_bg):
    """Apply header + row styling. Highlights duplicate names in yellow."""
    # Build headers
    headers = [c["header"] for c in columns]
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # Detect duplicate names
    name_key = next((c["key"] for c in columns if c["key"] in ("name", "memberName")), None)
    name_counts = {}
    if name_key:
        for r in row_data_list:
            n = (r.get(name_key) or "").lower().strip()
            if n:
                name_counts[n] = name_counts.get(n, 0) + 1

    for idx, row_data in enumerate(row_data_list):
        row_vals = [row_data.get(c["key"], "") for c in columns]
        ws.append(row_vals)
        n = (row_data.get(name_key) or "").lower().strip() if name_key else ""
        is_dup = name_key and n and name_counts.get(n, 0) > 1
        bg = DUPLICATE_BG if is_dup else get_row_bg(row_data, idx)
        row = ws[ws.max_row]
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border
        ws.row_dimensions[ws.max_row].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set column widths
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = col.get("width", 14)


# ═══════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════

# ── Auth ──────────────────────────────────────────────────────
@app.post("/api/auth/validate")
def auth_validate():
    body = request.get_json(force=True) or {}
    emp_id = (body.get("empId") or "").strip().lower()
    password = body.get("password") or ""
    if not emp_id:
        return jsonify({"error": "Employee ID is required"}), 400

    app_password = os.environ.get("APP_PASSWORD")
    if app_password:
        if not password:
            return jsonify({"error": "Password is required"}), 401
        if password != app_password:
            return jsonify({"error": "Invalid Employee ID or password"}), 401

    if emp_id in HARDCODED_CAREER_LEVELS:
        cl = HARDCODED_CAREER_LEVELS[emp_id]
        if cl >= 10:
            return jsonify({"error": "Access denied. Career level 9 and below only."}), 403
        return jsonify({"valid": True, "empId": emp_id, "careerLevel": cl})

    emp = next((e for e in employees
                if (e.get("id") or "").lower() == emp_id or (e.get("name") or "").lower() == emp_id), None)
    cl_raw = re.sub(r"[^0-9]", "", str(emp.get("careerLevel", "") if emp else ""))
    cl = int(cl_raw) if cl_raw else None

    if employees and emp and cl is not None and cl >= 10:
        return jsonify({"error": "Access denied. Career level 9 and below only."}), 403

    return jsonify({"valid": True, "empId": emp_id, "careerLevel": cl})


# ── Options ───────────────────────────────────────────────────
@app.get("/api/options/release-owners")
def get_release_owners():
    return jsonify(RELEASE_OWNERS)


@app.get("/api/options/career-levels")
def get_career_levels():
    return jsonify(CAREER_LEVELS)


@app.get("/api/options/employee-lookup")
def employee_lookup():
    q = (request.args.get("q") or "").strip().lower()
    if not q:
        return jsonify(None)
    match = next((e for e in employees
                  if (e.get("id") or "").lower() == q or (e.get("name") or "").lower() == q), None)
    return jsonify(match)


@app.get("/api/options/employees")
def get_employees():
    return jsonify(employees)


@app.get("/api/options/resource-upload-history")
def get_resource_upload_history():
    sorted_history = sorted(resource_upload_history, key=lambda x: x.get("uploadedAt", ""), reverse=True)
    return jsonify(sorted_history)


@app.post("/api/options/employees")
def post_employees():
    global employees
    body = request.get_json(force=True) or {}
    if isinstance(body, list):
        emp_list, uploaded_by, filename = body, "UNKNOWN", "unknown"
    else:
        emp_list = body.get("employees", [])
        uploaded_by = body.get("uploadedBy", "UNKNOWN")
        filename = body.get("filename", "unknown")
    if not isinstance(emp_list, list) or len(emp_list) == 0:
        return jsonify({"error": "Array of employees required"}), 400
    employees = emp_list
    resource_upload_history.append({
        "id": new_id(), "uploadedBy": uploaded_by, "filename": filename,
        "uploadedAt": now_iso(), "count": len(emp_list),
    })
    save_data()
    return jsonify({"success": True, "count": len(emp_list)})


# ── Entries ───────────────────────────────────────────────────
@app.get("/api/entries")
def get_entries():
    lst = [
        {"id": e["id"], "date": e["date"], "releaseOwner": e["releaseOwner"], "createdBy": e.get("createdBy", "")}
        for e in entries.values()
    ]
    lst.sort(key=lambda x: x["date"], reverse=True)
    return jsonify(lst)


@app.get("/api/entries/<eid>")
def get_entry(eid):
    e = entries.get(eid)
    if not e:
        return jsonify({"error": "Entry not found"}), 404
    return jsonify(e)


@app.post("/api/entries")
def create_entry():
    release_owner = request.form.get("releaseOwner") or (request.get_json(force=True) or {}).get("releaseOwner")
    date = request.form.get("date") or (request.get_json(force=True) or {}).get("date")
    created_by = request.form.get("createdBy") or (request.get_json(force=True) or {}).get("createdBy") or "UNKNOWN"

    if not release_owner or not date:
        return jsonify({"error": "Release Owner and Date are required"}), 400

    existing = next((e for e in entries.values() if e["date"] == date), None)
    if existing:
        return jsonify({"error": "Entry already exists for this date"}), 409

    sanity_file = None
    if "sanityFile" in request.files:
        f = request.files["sanityFile"]
        if f and f.filename and allowed_file(f.filename):
            fname = f"{int(datetime.now().timestamp() * 1000)}-{secure_filename(f.filename)}"
            f.save(UPLOADS_DIR / fname)
            sanity_file = fname

    eid = new_id()
    entries[eid] = {
        "id": eid, "releaseOwner": release_owner, "date": date,
        "sanityFile": sanity_file, "createdBy": created_by, "teams": [],
    }
    save_data()
    return jsonify(entries[eid]), 201


@app.put("/api/entries/<eid>")
def update_entry(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    body = request.get_json(force=True) or {}
    release_owner = body.get("releaseOwner")
    date = body.get("date")
    changed_by = (body.get("changedBy") or "UNKNOWN").strip()

    if eid not in changelogs:
        changelogs[eid] = []

    if release_owner is not None:
        if not release_owner.strip():
            return jsonify({"error": "Release Owner cannot be empty"}), 400
        if entry["releaseOwner"] != release_owner.strip():
            changelogs[eid].append({
                "id": new_id(), "type": "ownerChange",
                "oldValue": entry["releaseOwner"], "newValue": release_owner.strip(),
                "changedBy": changed_by, "timestamp": now_iso(),
            })
        entry["releaseOwner"] = release_owner.strip()

    if date is not None:
        if not date.strip():
            return jsonify({"error": "Date cannot be empty"}), 400
        dup = next((e for e in entries.values() if e["date"] == date.strip() and e["id"] != eid), None)
        if dup:
            return jsonify({"error": "Another entry already exists for this date"}), 409
        if entry["date"] != date.strip():
            changelogs[eid].append({
                "id": new_id(), "type": "dateChange",
                "oldValue": entry["date"], "newValue": date.strip(),
                "changedBy": changed_by, "timestamp": now_iso(),
            })
        entry["date"] = date.strip()

    save_data()
    return jsonify(entry)


@app.delete("/api/entries/<eid>")
def delete_entry(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    import copy
    deleted_items.append({
        "id": new_id(), "type": "entry", "data": copy.deepcopy(entry),
        "deletedAt": now_iso(),
        "parentInfo": {"date": entry["date"], "releaseOwner": entry["releaseOwner"]},
    })
    del entries[eid]
    save_data()
    return jsonify({"success": True})


# ── Teams ─────────────────────────────────────────────────────
@app.post("/api/entries/<eid>/teams")
def create_team(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    body = request.get_json(force=True) or {}
    team_name = body.get("teamName", "")
    lead_name = body.get("leadName", "")
    total_count = body.get("totalCount", 0)
    created_by = body.get("createdBy", "UNKNOWN")

    if not team_name or not lead_name:
        return jsonify({"error": "Team Name and Lead Name are required"}), 400
    if any(t["teamName"].lower() == team_name.lower() for t in entry["teams"]):
        return jsonify({"error": f'Team "{team_name}" already exists'}), 409

    team = {
        "id": new_id(), "teamName": team_name, "leadName": lead_name,
        "totalCount": int(total_count) if total_count else 0,
        "createdDate": entry["date"], "createdBy": created_by, "lineItems": [],
    }
    entry["teams"].append(team)
    save_data()
    return jsonify(team), 201


@app.put("/api/entries/<eid>/teams/<tid>")
def update_team(eid, tid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    idx = next((i for i, t in enumerate(entry["teams"]) if t["id"] == tid), -1)
    if idx == -1:
        return jsonify({"error": "Team not found"}), 404
    body = request.get_json(force=True) or {}
    entry["teams"][idx] = {**entry["teams"][idx], **body, "id": tid}
    save_data()
    return jsonify(entry["teams"][idx])


@app.delete("/api/entries/<eid>/teams/<tid>")
def delete_team(eid, tid):
    import copy
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if team:
        deleted_items.append({
            "id": new_id(), "type": "team", "data": copy.deepcopy(team),
            "deletedAt": now_iso(),
            "parentInfo": {"entryId": eid, "date": entry["date"], "teamName": team["teamName"]},
        })
    entry["teams"] = [t for t in entry["teams"] if t["id"] != tid]
    save_data()
    return jsonify({"success": True})


# ── Line Items ────────────────────────────────────────────────
@app.post("/api/entries/<eid>/teams/<tid>/line-items")
def create_line_item(eid, tid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404
    body = request.get_json(force=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if any(li["name"].lower() == name.lower() for li in team["lineItems"]):
        return jsonify({"error": f'"{name}" already exists in this team'}), 409
    item = {
        "id": new_id(), "name": name,
        "careerLevel": body.get("careerLevel", ""),
        "allowanceCompoff": body.get("allowanceCompoff", "Compoff"),
        "time": body.get("time", ""),
    }
    team["lineItems"].append(item)
    save_data()
    return jsonify(item), 201


@app.post("/api/entries/<eid>/teams/<tid>/bulk-line-items")
def bulk_line_items(eid, tid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404
    body = request.get_json(force=True) or {}
    items = body.get("items", [])
    if not isinstance(items, list):
        return jsonify({"error": "items array required"}), 400

    added, updated = [], []
    for item in items:
        name = (item.get("name") or "").strip()
        if not name:
            continue
        cl = item.get("careerLevel", "")
        sv = item.get("supervisor", "")
        ac = item.get("allowanceCompoff", "Compoff")
        idx = next((i for i, li in enumerate(team["lineItems"]) if li["name"].lower() == name.lower()), -1)
        if idx != -1:
            team["lineItems"][idx] = {
                **team["lineItems"][idx],
                "name": name,
                "careerLevel": cl or team["lineItems"][idx].get("careerLevel", ""),
                "supervisor": sv or team["lineItems"][idx].get("supervisor", ""),
                "allowanceCompoff": ac,
            }
            updated.append(team["lineItems"][idx])
        else:
            new_item = {"id": new_id(), "name": name, "careerLevel": cl, "supervisor": sv, "allowanceCompoff": ac}
            team["lineItems"].append(new_item)
            added.append(new_item)

    save_data()
    return jsonify({"added": added, "updated": updated, "all": team["lineItems"]}), 201


@app.put("/api/entries/<eid>/teams/<tid>/line-items/<lid>")
def update_line_item(eid, tid, lid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404
    idx = next((i for i, li in enumerate(team["lineItems"]) if li["id"] == lid), -1)
    if idx == -1:
        return jsonify({"error": "Item not found"}), 404
    body = request.get_json(force=True) or {}
    team["lineItems"][idx] = {**team["lineItems"][idx], **body, "id": lid}
    save_data()
    return jsonify(team["lineItems"][idx])


@app.delete("/api/entries/<eid>/teams/<tid>/line-items/<lid>")
def delete_line_item(eid, tid, lid):
    import copy
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404
    item = next((li for li in team["lineItems"] if li["id"] == lid), None)
    if item:
        deleted_items.append({
            "id": new_id(), "type": "lineItem", "data": copy.deepcopy(item),
            "deletedAt": now_iso(),
            "parentInfo": {
                "entryId": eid, "teamId": tid, "date": entry["date"],
                "teamName": team["teamName"], "memberName": item["name"],
            },
        })
    team["lineItems"] = [li for li in team["lineItems"] if li["id"] != lid]
    save_data()
    return jsonify({"success": True})


@app.put("/api/entries/<eid>/teams/<tid>/line-items")
def replace_line_items(eid, tid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404
    body = request.get_json(force=True) or {}
    items = body.get("items", [])
    if not isinstance(items, list):
        return jsonify({"error": "items array required"}), 400

    prev_count = len(team["lineItems"])
    team["lineItems"] = [
        {
            "id": li.get("id") or new_id(),
            "name": li.get("name", ""),
            "careerLevel": li.get("careerLevel", ""),
            "supervisor": li.get("supervisor", ""),
            "allowanceCompoff": li.get("allowanceCompoff", "Compoff"),
            "time": li.get("time", ""),
            "notes": li.get("notes", ""),
        }
        for li in items
    ]

    changed_by = (body.get("changedBy") or "UNKNOWN").strip()
    if eid not in changelogs:
        changelogs[eid] = []
    changelogs[eid].append({
        "id": new_id(), "type": "teamSave",
        "oldValue": str(prev_count), "newValue": team["teamName"],
        "changedBy": changed_by, "teamName": team["teamName"], "leadName": team["leadName"],
        "timestamp": now_iso(),
    })
    save_data()
    return jsonify(team["lineItems"])


# ── Bulk Teams Upload ─────────────────────────────────────────
@app.post("/api/entries/<eid>/bulk-teams")
def bulk_teams(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    body = request.get_json(force=True) or {}
    teams = body.get("teams", [])
    if not isinstance(teams, list):
        return jsonify({"error": "teams array required"}), 400
    created_by = body.get("createdBy", "UNKNOWN")
    added, skipped = [], []
    for t in teams:
        tn = (t.get("teamName") or "").strip()
        if not tn:
            continue
        ln = (t.get("leadName") or "").strip() or "TBD"
        if any(et["teamName"].lower() == tn.lower() for et in entry["teams"]):
            skipped.append(tn)
            continue
        team = {
            "id": new_id(), "teamName": tn, "leadName": ln, "totalCount": 0,
            "createdDate": entry["date"], "createdBy": created_by, "lineItems": [],
        }
        entry["teams"].append(team)
        added.append(team)
    save_data()
    return jsonify({"added": added, "skipped": skipped}), 201


# ── Deleted Items ─────────────────────────────────────────────
@app.get("/api/deleted-items")
def get_deleted_items():
    return jsonify(sorted(deleted_items, key=lambda x: x.get("deletedAt", ""), reverse=True))


@app.post("/api/deleted-items/<did>/recover")
def recover_deleted(did):
    import copy
    idx = next((i for i, d in enumerate(deleted_items) if d["id"] == did), -1)
    if idx == -1:
        return jsonify({"error": "Deleted item not found"}), 404
    item = deleted_items[idx]
    recovered = False

    if item["type"] == "entry":
        entries[item["data"]["id"]] = copy.deepcopy(item["data"])
        recovered = True
    elif item["type"] == "team":
        parent = entries.get(item["parentInfo"]["entryId"])
        if not parent:
            return jsonify({"error": "Parent entry no longer exists. Recover the entry first."}), 400
        parent["teams"].append(copy.deepcopy(item["data"]))
        recovered = True
    elif item["type"] == "lineItem":
        parent = entries.get(item["parentInfo"]["entryId"])
        if not parent:
            return jsonify({"error": "Parent entry no longer exists."}), 400
        team = next((t for t in parent["teams"] if t["id"] == item["parentInfo"]["teamId"]), None)
        if not team:
            return jsonify({"error": "Parent team no longer exists. Recover the team first."}), 400
        team["lineItems"].append(copy.deepcopy(item["data"]))
        recovered = True

    if recovered:
        deleted_items.pop(idx)
        save_data()
        return jsonify({"success": True})
    return jsonify({"error": "Could not recover item"}), 400


@app.delete("/api/deleted-items/<did>")
def delete_deleted_item(did):
    idx = next((i for i, d in enumerate(deleted_items) if d["id"] == did), -1)
    if idx == -1:
        return jsonify({"error": "Deleted item not found"}), 404
    deleted_items.pop(idx)
    save_data()
    return jsonify({"success": True})


@app.delete("/api/deleted-items")
def clear_deleted_items():
    global deleted_items
    deleted_items = []
    save_data()
    return jsonify({"success": True})


# ── Sanity File Upload ────────────────────────────────────────
@app.post("/api/entries/<eid>/sanity-upload")
def sanity_upload(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    if "sanityFile" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["sanityFile"]
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded"}), 400
    changed_by = (request.form.get("changedBy") or "UNKNOWN").strip()
    fname = f"{int(datetime.now().timestamp() * 1000)}-{secure_filename(f.filename)}"
    f.save(UPLOADS_DIR / fname)
    if eid not in changelogs:
        changelogs[eid] = []
    changelogs[eid].append({
        "id": new_id(), "type": "sanityUpload",
        "oldValue": entry.get("sanityFile", ""), "newValue": fname,
        "changedBy": changed_by, "timestamp": now_iso(),
    })
    entry["sanityFile"] = fname
    save_data()
    return jsonify({"sanityFile": fname})


# ── Last Updated ──────────────────────────────────────────────
@app.get("/api/last-updated")
def last_updated():
    lead_id = (request.args.get("leadId") or "").strip().lower()
    all_events = []
    for entry_id, logs in changelogs.items():
        entry = entries.get(entry_id)
        if not entry:
            continue
        for h in logs:
            if h.get("type") != "teamSave":
                continue
            if lead_id and (h.get("leadName") or "").lower() != lead_id:
                continue
            all_events.append({
                "id": h["id"], "entryId": entry_id, "entryDate": entry["date"],
                "teamName": h.get("teamName") or h.get("newValue", ""),
                "leadName": h.get("leadName", ""), "changedBy": h.get("changedBy", ""),
                "timestamp": h.get("timestamp", ""),
            })
    all_events.sort(key=lambda x: x["timestamp"], reverse=True)
    return jsonify(all_events[:5])


# ── Change History ────────────────────────────────────────────
@app.get("/api/entries/<eid>/history")
def get_history(eid):
    if eid not in entries:
        return jsonify({"error": "Entry not found"}), 404
    changelog = changelogs.get(eid, [])
    normalized = [
        {
            "id": h["id"], "type": h["type"],
            "oldValue": h.get("oldValue", h.get("old", "")),
            "newValue": h.get("newValue", h.get("new", "")),
            "changedBy": h.get("changedBy", ""),
            "timestamp": h.get("timestamp") or h.get("date") or now_iso(),
        }
        for h in changelog
    ]
    return jsonify(normalized)


@app.delete("/api/entries/<eid>/history")
def clear_history(eid):
    if eid not in entries:
        return jsonify({"error": "Entry not found"}), 404
    changelogs[eid] = []
    save_data()
    return jsonify({"success": True})


@app.delete("/api/entries/<eid>/history/by-date/<target_date>")
def delete_history_by_date(eid, target_date):
    if eid not in entries:
        return jsonify({"error": "Entry not found"}), 404
    if eid not in changelogs:
        changelogs[eid] = []
        return jsonify({"success": True, "deleted": 0})
    before = len(changelogs[eid])
    changelogs[eid] = [
        h for h in changelogs[eid]
        if (h.get("timestamp") or h.get("date") or "")[:10] != target_date
    ]
    deleted_count = before - len(changelogs[eid])
    save_data()
    return jsonify({"success": True, "deleted": deleted_count})


# ── Excel Exports ─────────────────────────────────────────────
ALLOWANCE_BG = "FFD1FAE5"
COMPOFF_BG = "FFFEF3C7"
ALT_COLORS = ["FFEEF2FF", "FFF5F3FF"]


@app.get("/api/entries/<eid>/export")
def export_entry(eid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Release"

    columns = [
        {"header": "Release Owner", "key": "releaseOwner", "width": 22},
        {"header": "Release Date", "key": "releaseDate", "width": 14},
        {"header": "Team Name", "key": "teamName", "width": 22},
        {"header": "Lead Name", "key": "leadName", "width": 22},
        {"header": "Member Name", "key": "memberName", "width": 28},
        {"header": "Career Level", "key": "careerLevel", "width": 14},
        {"header": "Login Time", "key": "loginTime", "width": 14},
        {"header": "Logout Time", "key": "logoutTime", "width": 14},
        {"header": "Total Hours", "key": "totalHours", "width": 14},
        {"header": "Compoff/Allowance", "key": "type", "width": 20},
        {"header": "Created By", "key": "createdBy", "width": 22},
    ]

    row_data_list = []
    for t in entry["teams"]:
        if not t["lineItems"]:
            row_data_list.append({
                "releaseOwner": entry["releaseOwner"], "releaseDate": entry["date"],
                "teamName": t["teamName"], "leadName": t["leadName"],
                "memberName": "", "careerLevel": "", "loginTime": "", "logoutTime": "",
                "totalHours": "", "type": "", "createdBy": t.get("createdBy", ""),
                "_type": "",
            })
        else:
            for i, li in enumerate(t["lineItems"]):
                parts = [p.strip() for p in (li.get("time") or "").split("-", 1)]
                emp_info = lookup_employee_info(li)
                enriched_li = {**li, "empId": emp_info["empId"], "level": emp_info["level"]}
                row_data_list.append({
                    "releaseOwner": entry["releaseOwner"] if i == 0 else "",
                    "releaseDate": entry["date"] if i == 0 else "",
                    "teamName": t["teamName"] if i == 0 else "",
                    "leadName": t["leadName"] if i == 0 else "",
                    "memberName": (li["name"] + (f" ({emp_info['empId']})" if emp_info["empId"] else "")) if li.get("name") else "",
                    "careerLevel": get_numeric_level(enriched_li),
                    "loginTime": parts[0] if len(parts) > 0 else "",
                    "logoutTime": parts[1] if len(parts) > 1 else "",
                    "totalHours": calc_total_hours(li.get("time", "")),
                    "type": li.get("allowanceCompoff", ""),
                    "createdBy": t.get("createdBy", "") if i == 0 else "",
                    "_type": li.get("allowanceCompoff", ""),
                })

    def get_bg(row, idx):
        t = row.get("_type", "")
        if t == "Allowance":
            return ALLOWANCE_BG
        if t == "Compoff":
            return COMPOFF_BG
        return ALT_COLORS[idx % 2]

    style_sheet(ws, columns, row_data_list, get_bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Release-{entry['date']}_{file_timestamp()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/entries/<eid>/teams/<tid>/export")
def export_team(eid, tid):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    team = next((t for t in entry["teams"] if t["id"] == tid), None)
    if not team:
        return jsonify({"error": "Team not found"}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (team["teamName"] or "Team")[:31]

    columns = [
        {"header": "Name", "key": "name", "width": 28},
        {"header": "Career Level", "key": "careerLevel", "width": 14},
        {"header": "Supervisor", "key": "supervisor", "width": 22},
        {"header": "Login Time", "key": "loginTime", "width": 14},
        {"header": "Logout Time", "key": "logoutTime", "width": 14},
        {"header": "Total Hours", "key": "totalHours", "width": 14},
        {"header": "Compoff/Allowance", "key": "type", "width": 20},
        {"header": "Lead", "key": "lead", "width": 22},
    ]

    row_data_list = []
    for li in team["lineItems"]:
        parts = [p.strip() for p in (li.get("time") or "").split("-", 1)]
        emp_info = lookup_employee_info(li)
        enriched_li = {**li, "empId": emp_info["empId"], "level": emp_info["level"]}
        row_data_list.append({
            "name": (li.get("name", "") + (f" ({emp_info['empId']})" if emp_info["empId"] else "")) if li.get("name") else "",
            "careerLevel": get_numeric_level(enriched_li),
            "supervisor": li.get("supervisor", ""),
            "loginTime": parts[0] if len(parts) > 0 else "",
            "logoutTime": parts[1] if len(parts) > 1 else "",
            "totalHours": calc_total_hours(li.get("time", "")),
            "type": li.get("allowanceCompoff", ""),
            "lead": team["leadName"],
            "_type": li.get("allowanceCompoff", ""),
        })

    def get_bg(row, idx):
        t = row.get("_type", "")
        if t == "Allowance":
            return ALLOWANCE_BG
        if t == "Compoff":
            return COMPOFF_BG
        return ALT_COLORS[idx % 2]

    style_sheet(ws, columns, row_data_list, get_bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", title_case(team["teamName"]))
    filename = f"{safe_name}-{entry['date']}_{file_timestamp()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/entries/<eid>/dialog-export/<export_type>")
def dialog_export(eid, export_type):
    entry = entries.get(eid)
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    teams = entry.get("teams", [])
    all_items = [
        {**li, "teamName": t["teamName"], "leadName": t["leadName"]}
        for t in teams for li in (t.get("lineItems") or [])
    ]
    allowance_items = [li for li in all_items if li.get("allowanceCompoff") == "Allowance"]
    compoff_items = [li for li in all_items if li.get("allowanceCompoff") == "Compoff"]

    if export_type == "allowance":
        sheet_name = "Shift Allowance"
        columns = [
            {"header": "#", "key": "num", "width": 5},
            {"header": "Name", "key": "name", "width": 28},
            {"header": "Career Level", "key": "careerLevel", "width": 14},
            {"header": "Supervisor", "key": "supervisor", "width": 22},
            {"header": "Team", "key": "team", "width": 20},
            {"header": "Lead", "key": "lead", "width": 22},
            {"header": "Login Time", "key": "loginTime", "width": 14},
            {"header": "Logout Time", "key": "logoutTime", "width": 14},
            {"header": "Total Hours", "key": "totalHours", "width": 14},
            {"header": "Support type to be charged in myte", "key": "supportType", "width": 38},
        ]
        row_data_list = []
        for i, li in enumerate(allowance_items):
            parts = [p.strip() for p in (li.get("time") or "").split("-", 1)]
            hrs = calc_total_hours(li.get("time", ""))
            emp_info = lookup_employee_info(li)
            enriched_li = {**li, "empId": emp_info["empId"], "level": emp_info["level"]}
            row_data_list.append({
                "num": i + 1, "name": (li["name"] + (f" ({emp_info['empId']})" if emp_info["empId"] else "")) if li.get("name") else "", "careerLevel": get_numeric_level(enriched_li),
                "supervisor": li.get("supervisor", ""), "team": li["teamName"], "lead": li["leadName"],
                "loginTime": parts[0] if len(parts) > 0 else "",
                "logoutTime": parts[1] if len(parts) > 1 else "",
                "totalHours": hrs, "supportType": get_support_type(hrs),
            })
        get_bg = lambda r, i: ALLOWANCE_BG

    elif export_type == "compoff":
        sheet_name = "Comp-off"
        columns = [
            {"header": "#", "key": "num", "width": 5},
            {"header": "Name", "key": "name", "width": 28},
            {"header": "Career Level", "key": "careerLevel", "width": 14},
            {"header": "Supervisor", "key": "supervisor", "width": 22},
            {"header": "Team", "key": "team", "width": 20},
            {"header": "Lead", "key": "lead", "width": 22},
            {"header": "Login Time", "key": "loginTime", "width": 14},
            {"header": "Logout Time", "key": "logoutTime", "width": 14},
            {"header": "Total Hours", "key": "totalHours", "width": 14},
        ]
        row_data_list = []
        for i, li in enumerate(compoff_items):
            parts = [p.strip() for p in (li.get("time") or "").split("-", 1)]
            hrs = calc_total_hours(li.get("time", ""))
            emp_info = lookup_employee_info(li)
            enriched_li = {**li, "empId": emp_info["empId"], "level": emp_info["level"]}
            row_data_list.append({
                "num": i + 1, "name": (li["name"] + (f" ({emp_info['empId']})" if emp_info["empId"] else "")) if li.get("name") else "", "careerLevel": get_numeric_level(enriched_li),
                "supervisor": li.get("supervisor", ""), "team": li["teamName"], "lead": li["leadName"],
                "loginTime": parts[0] if len(parts) > 0 else "",
                "logoutTime": parts[1] if len(parts) > 1 else "",
                "totalHours": hrs,
            })
        get_bg = lambda r, i: COMPOFF_BG

    elif export_type == "members":
        sheet_name = "All Members"
        columns = [
            {"header": "#", "key": "num", "width": 5},
            {"header": "Name", "key": "name", "width": 28},
            {"header": "Career Level", "key": "careerLevel", "width": 14},
            {"header": "Supervisor", "key": "supervisor", "width": 22},
            {"header": "Team", "key": "team", "width": 20},
            {"header": "Lead", "key": "lead", "width": 22},
            {"header": "Login Time", "key": "loginTime", "width": 14},
            {"header": "Logout Time", "key": "logoutTime", "width": 14},
            {"header": "Total Hours", "key": "totalHours", "width": 14},
            {"header": "Type", "key": "type", "width": 16},
        ]
        row_data_list = []
        for i, li in enumerate(all_items):
            parts = [p.strip() for p in (li.get("time") or "").split("-", 1)]
            hrs = calc_total_hours(li.get("time", ""))
            emp_info = lookup_employee_info(li)
            enriched_li = {**li, "empId": emp_info["empId"], "level": emp_info["level"]}
            row_data_list.append({
                "num": i + 1, "name": (li["name"] + (f" ({emp_info['empId']})" if emp_info["empId"] else "")) if li.get("name") else "", "careerLevel": get_numeric_level(enriched_li),
                "supervisor": li.get("supervisor", ""), "team": li["teamName"], "lead": li["leadName"],
                "loginTime": parts[0] if len(parts) > 0 else "",
                "logoutTime": parts[1] if len(parts) > 1 else "",
                "totalHours": hrs, "type": li.get("allowanceCompoff", ""),
                "_type": li.get("allowanceCompoff", ""),
            })

        def get_bg(row, idx):
            t = row.get("_type", "")
            if t == "Allowance":
                return ALLOWANCE_BG
            if t == "Compoff":
                return COMPOFF_BG
            return ALT_COLORS[idx % 2]

    elif export_type == "teams":
        sheet_name = "Teams"
        columns = [
            {"header": "#", "key": "num", "width": 5},
            {"header": "Team Name", "key": "teamName", "width": 28},
            {"header": "Lead", "key": "lead", "width": 22},
            {"header": "Members", "key": "members", "width": 12},
        ]
        row_data_list = [
            {"num": i + 1, "teamName": t["teamName"], "lead": t["leadName"],
             "members": len(t.get("lineItems", []))}
            for i, t in enumerate(teams)
        ]
        get_bg = lambda r, i: ALT_COLORS[i % 2]
    else:
        return jsonify({"error": "Unknown export type"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    style_sheet(ws, columns, row_data_list, get_bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    label_map = {"allowance": "Shift Allowance", "compoff": "Compoff",
                 "members": "All Members", "teams": "Teams"}
    label = label_map.get(export_type, export_type)
    filename = f"{label}-{entry['date']}_{file_timestamp()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Static file serving ───────────────────────────────────────
@app.get("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(UPLOADS_DIR, filename)


if FRONTEND_BUILD.exists():
    @app.get("/")
    def serve_index():
        return send_from_directory(str(FRONTEND_BUILD), "index.html")

    @app.get("/<path:path>")
    def serve_static(path):
        full = FRONTEND_BUILD / path
        if full.exists() and full.is_file():
            return send_from_directory(str(FRONTEND_BUILD), path)
        return send_from_directory(str(FRONTEND_BUILD), "index.html")


# ── Startup ───────────────────────────────────────────────────
def start():
    global entries, deleted_items, changelogs, employees, resource_upload_history

    # Load from Azure Blob first, fall back to local data.json
    azure_data = _load_from_azure()
    file_data = load_data_from_file()
    data = azure_data or file_data or {}

    entries = data.get("entries", {})
    deleted_items = data.get("deletedItems", [])
    changelogs = data.get("changelogs", {})
    employees = data.get("employees") or []
    resource_upload_history = data.get("resourceUploadHistory", [])

    print(f"Loaded {len(entries)} entries from {'Azure' if azure_data else 'local file' if file_data else 'scratch'}.")
    print(f" * Running on http://localhost:{PORT}")
    app.run(host="0.0.0.0", port=PORT, debug=False)


if __name__ == "__main__":
    start()
